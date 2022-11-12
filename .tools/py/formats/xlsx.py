from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from math import ceil
from shared.common import *

def calc_row_height(text, font_size, col_width, line_height=16.0):
    line_count = 0
    for line in text.split(U'\n'):
        half_width_count = 0
        for c in line:
            n = ord(c)
            if n < 0x80:
                if n < 0x20:
                    continue
                half_width_count += 1
            else:
                half_width_count += 1
                try:
                    test_c = c.encode('cp932')
                    if len(test_c) > 1:
                        half_width_count += 1
                except UnicodeEncodeError:
                    pass

        char_per_line = int(ceil(col_width / (font_size * 0.1)) + 6)  # stupid calculator

        if half_width_count == 0:
            half_width_count = 1

        quotient, remainder = divmod(half_width_count,
                                     char_per_line)

        line_count += quotient
        if remainder > 0:
            line_count += 1

    return max([line_height, line_height * line_count])


def write_xlsx(out_path, header, rows):
    font_size = 12
    text_col_width = 100.0
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active
    ws.title = get_filename(out_path, False)
    ws.append(header)

    # header row
    hfont = Font(name='Calibri',
                 size=font_size + 2,
                 bold=True)
    h_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    header_row = ws.row_dimensions[1]
    header_row.height = 20
    header_row.font = hfont
    header_row.alignment = h_alignment

    id_col = ws.column_dimensions['A']
    jp_col = ws.column_dimensions['B']
    vn_col = ws.column_dimensions['C']
    notes_col = ws.column_dimensions['D']

    alignment = Alignment(wrap_text=True, vertical='top')
    font = Font(name='Calibri',
                size=font_size,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')

    # update style
    for i, col in enumerate([id_col, jp_col, vn_col, notes_col]):
        if i == 0:
            col.width = 20.0
        else:
            col.width = text_col_width
        col.bestFit = True
        col.auto_size = True
        col.alignment = alignment
        col.font = font

    # write row data
    row_idx = 2  # header row = 1
    for row in rows:
        ws.append(row)
        ws.row_dimensions[row_idx].height = calc_row_height(row[1], font_size, text_col_width)
        row_idx += 1

    wb.save(out_path)
    pass


def load_xlsx_trans_file(xlsx_path):
    sheet_name = get_filename(xlsx_path, False)
    wb = load_workbook(xlsx_path, True)

    sheet = wb[sheet_name]

    rows_it = sheet.rows
    next(rows_it)  # skip header
    res = []

    for row in rows_it:
        line_id_cell = row[0]
        jp_text_cell = row[1]
        vn_text_cell = row[2]

        line_id = line_id_cell.value
        jp_text = jp_text_cell.value
        vn_text = vn_text_cell.value

        res.append([line_id, jp_text, vn_text])

    return res

    pass