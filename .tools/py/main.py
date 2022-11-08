# coding=utf-8
import glob
import csv
import sys
import argparse

from bs4 import BeautifulSoup, Tag, NavigableString, PageElement
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from typing import List
from shared.common import *
from math import ceil
from io import StringIO

CSV_FORMAT = 0
TXT_FORMAT = 1
XLSX_FORMAT = 2

TXT_ID = 0
TXT_JP = 1
TXT_VN = 2
TXT_IDENTIFY = {U'ID': TXT_ID, U'JP': TXT_JP, U'VN': TXT_VN}
TXT_EX_END_FLAG = U'=--='

TITLE_TAG = 'title'
META_TAG = 'meta'
TRANS_TAG = 'trans'

META_NAME = [U'description', U'keywords']
META_PROP = [U'og:title', U'og:site_name', U'og:description']

PROPERTY = 'property'
CONTENT = 'content'
NAME = 'name'

CSV_HEADER = [U'ID', U'JP', U'VN', U'Notes']


def dump_head_text(head_elem):
    # type: (Tag) -> List

    result = []
    tags = head_elem.find_all([TITLE_TAG, META_TAG])  # type: List[Tag]

    for tag in tags:
        if tag.name == META_TAG:
            name_attr = tag.attrs.get(NAME)
            prop_attr = tag.attrs.get(PROPERTY)
            content = tag.attrs.get(CONTENT)

            if name_attr is not None and name_attr in META_NAME:
                result.append([U'{}_{}'.format(META_TAG, name_attr),
                               content,
                               U'',
                               U''])
            elif prop_attr is not None and prop_attr in META_PROP:
                result.append([U'{}_{}'.format(META_TAG, prop_attr),
                               content,
                               U'',
                               U''])

        elif tag.name == TITLE_TAG:
            result.append([TITLE_TAG, tag.text, U''])

    return result


def dump_body_text(body_elem):
    # type: (Tag) -> List

    result = []
    str_idx = 0

    prev_string = U''
    string_elem_idxs = []
    strings_it = iter(body_elem.strings)

    for string in strings_it:  # type: NavigableString
        if string.strip():
            parent_elem = string.parent
            parent_class = parent_elem.attrs.get('class')
            is_voice_line = False
            curr_voice_class = None

            if parent_class:
                if U'jp-play' in parent_class:
                    is_voice_line = True
                    curr_voice_class = U'jp-play'
                elif U'jp-pause' in parent_class:
                    is_voice_line = True
                    curr_voice_class = U'jp-pause'
                elif U'novel_copyright' in parent_class:
                    continue

            text_line_elems = []
            text_line = unicode(string)
            text_line_elems.append(string)

            next_str = string.next_sibling  # type: Tag | PageElement
            while next_str is not None:
                if next_str.name == 'ruby':
                    text_line += unicode(next_str)
                    text_line_elems.append(next_str)
                    next_str = next_str.next_sibling

                    next(strings_it)  # text
                    next(strings_it)  # furigana

                elif type(next_str) is NavigableString:
                    if next_str.strip():
                        text_line += unicode(next_str)
                        text_line_elems.append(next_str)
                        next_str = next_str.next_sibling
                        next(strings_it)

                else:
                    break

            if prev_string == text_line and curr_voice_class == U'jp-pause':
                # duplicate text voice -> map to prev idx
                string_elem_idxs.append([str_idx - 1, text_line_elems])
                continue
            else:
                string_elem_idxs.append([str_idx, text_line_elems])

                result.append([unicode(str_idx), text_line, U'', U''])
                prev_string = text_line
                str_idx += 1

    # replace text elem with trans tag
    for idx, strings in string_elem_idxs:
        trans_tag = BeautifulSoup("<{} idx=\"{}\">{}</{}>"
                                  .format(TRANS_TAG, idx, U''
                                          .join([unicode(s) for s in strings]).encode('utf8'), TRANS_TAG),
                                  'html.parser')

        while len(strings) > 1:
            string = strings.pop()
            string.replace_with("")

        strings[0].replace_with(trans_tag)

    return result


def write_csv(out_path, rows):
    writer = csv.writer(
        open(out_path, 'wb'),
        delimiter=',', quotechar='"',
        quoting=csv.QUOTE_MINIMAL)

    for row in rows:
        writer.writerow([c.encode('utf8') for c in row])


def calc_row_height(text, font_size, col_width, line_height=16.0):
    line_count = 0
    for line in text.split(U'\n'):
        half_width_count = 0
        for c in line:
            n = ord(c)
            if n < 0x80:
                if n < 0x20:
                    continue
                half_width_count += 1
            else:
                half_width_count += 1
                try:
                    test_c = c.encode('cp932')
                    if len(test_c) > 1:
                        half_width_count += 1
                except UnicodeEncodeError:
                    pass

        char_per_line = int(ceil(col_width / (font_size * 0.1)) + 6)  # stupid calculator

        if half_width_count == 0:
            half_width_count = 1

        quotient, remainder = divmod(half_width_count,
                                     char_per_line)

        line_count += quotient
        if remainder > 0:
            line_count += 1

    return max([line_height, line_height * line_count])


def write_xlsx(out_path, header, rows):
    font_size = 12
    text_col_width = 100.0
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active
    ws.title = get_filename(out_path, False)
    ws.append(header)

    # header row
    hfont = Font(name='Calibri',
                 size=font_size + 2,
                 bold=True)
    h_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    header_row = ws.row_dimensions[1]
    header_row.height = 20
    header_row.font = hfont
    header_row.alignment = h_alignment

    id_col = ws.column_dimensions['A']
    jp_col = ws.column_dimensions['B']
    vn_col = ws.column_dimensions['C']
    notes_col = ws.column_dimensions['D']

    alignment = Alignment(wrap_text=True, vertical='top')
    font = Font(name='Calibri',
                size=font_size,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')

    # update style
    for i, col in enumerate([id_col, jp_col, vn_col, notes_col]):
        if i == 0:
            col.width = 20.0
        else:
            col.width = text_col_width
        col.bestFit = True
        col.auto_size = True
        col.alignment = alignment
        col.font = font

    # write row data
    row_idx = 2  # header row = 1
    for row in rows:
        ws.append(row)
        ws.row_dimensions[row_idx].height = calc_row_height(row[1], font_size, text_col_width)
        row_idx += 1

    wb.save(out_path)
    pass


def write_txt(out_path, rows):
    fi = open(out_path, 'wb')
    dict_for_toml = {}
    for row in rows:
        line_id = row[0]
        jp_text = row[1]

        fi.write(U'[ID]:{}\n'.format(line_id).encode('utf8'))
        fi.write(U'[JP]:\n{}\n'.format(jp_text).encode('utf8'))
        fi.write('[VN]:\n\n')
        fi.write('{}========================================\n'.format(TXT_EX_END_FLAG))

    fi.close()
    pass


def txt_clean_line(line, strip_char=None):
    ret = line.split(U'//', 1)[0]
    ret = ret.strip(strip_char)
    return ret


def txt_get_tag(fi):
    # type: (StringIO) -> [int, int]

    line_pos = fi.tell()
    line = fi.readline()

    # get identify tag
    identify_tag_str = None
    data_pos = -1
    while line != '':
        line_clean = txt_clean_line(line)
        open_pos = line.find(U'[')
        if open_pos != -1 and line_clean[0] == U'[':
            close_pos = line.find(U']')
            if close_pos > open_pos:
                data_pos = line_pos + close_pos + 2  # skip :
                identify_tag_str = line_clean[open_pos+1:close_pos]
                break
            else:
                print(U'[ERROR]: Close tag "]" not found: {}'.format(line_clean))
                raise ValueError()
        elif len(line_clean) > 0:
            return None
            # print(U'[ERROR]: Invalid declare tag: {}'.format(line_clean))
            # raise ValueError()

        line_pos = fi.tell()
        line = fi.readline()

    if identify_tag_str is None:
        return None

    identify_tag = TXT_IDENTIFY.get(identify_tag_str.upper())
    if identify_tag is None:
        return None
        # print(U'[ERROR]: Invalid tag: {}'.format(identify_tag_str))
        # raise ValueError()

    return [identify_tag, data_pos]
    pass


def txt_get_text_data(fi, data_start_pos):
    # type: (StringIO, int) -> unicode

    text_arr = []
    fi.seek(data_start_pos)
    first_char = fi.read(1)
    if first_char != U'\n':
        text_arr.append(first_char)

    line_pos = fi.tell()
    line = fi.readline()
    while line != U'':
        line_clean = txt_clean_line(line)
        if line_clean.startswith(U'['):
            tmp_pos = fi.tell()
            fi.seek(line_pos)
            # try to detect tag
            tag_info = txt_get_tag(fi)
            if tag_info is not None:
                # is tag -> go back and return
                fi.seek(line_pos)
                break

            fi.seek(tmp_pos)

        if line_clean.startswith(TXT_EX_END_FLAG):
            break

        text_arr.append(txt_clean_line(line, U''))  # clean comment only
        line_pos = fi.tell()
        line = fi.readline()

    res = U''.join(text_arr)
    if txt_clean_line(res) == U'':
        return U''

    if res[-1] == U'\n':
        res = res[:-1]

    return res.replace(U'\\n', U'<br/>')


def load_txt_trans_file(txt_path):
    fi = StringIO(open(txt_path, 'rb').read().replace('\r\n', '\n').decode('utf8'))
    dict_res = {}

    while True:
        start_pos = fi.tell()
        tag_info = txt_get_tag(fi)
        if tag_info is None:
            # check is eof
            fi.seek(start_pos)
            line = fi.readline()
            if line != U'':
                print(U'[ERROR]: Tag invalid: {}'.format(line))
                raise ValueError
            break

        identify_tag, data_pos = tag_info
        fi.seek(data_pos)

        if identify_tag != TXT_ID:
            print(U'[ERROR]: Require [ID] tag: {}'.format(fi.readline()))
            raise ValueError

        line_id = txt_clean_line(fi.readline())

        identify_tag, data_pos = txt_get_tag(fi)  # jp
        jp_text = txt_get_text_data(fi, data_pos)

        identify_tag, data_pos = txt_get_tag(fi)  # vn
        vn_text = txt_get_text_data(fi, data_pos)

        dict_res[line_id] = [jp_text, vn_text]

    return dict_res


def html_to_text(html_folder, out_folder, out_format=XLSX_FORMAT):
    html_paths = glob.glob(U'{}/*.html'.format(html_folder))

    # for import
    html_out_folder = U'{}/html'.format(out_folder)

    if is_folder(out_folder) is False:
        make_dirs(out_folder)
    if is_folder(html_out_folder) is False:
        make_dirs(html_out_folder)

    for html_path in html_paths:
        print(html_path)

        html_text = open(html_path, 'rb').read()
        dom = BeautifulSoup(html_text, features="html.parser")

        head_rows = dump_head_text(dom.head)
        body_rows = dump_body_text(dom.body)

        filename = get_filename(html_path, False)

        if out_format == CSV_FORMAT:
            out_path = U'{}/{}.csv'.format(out_folder, filename)
            write_csv(out_path, [CSV_HEADER] + head_rows + body_rows)
        elif out_format == XLSX_FORMAT:
            out_path = U'{}/{}.xlsx'.format(out_folder, filename)
            write_xlsx(out_path, CSV_HEADER, head_rows + body_rows)
        elif out_format == TXT_FORMAT:
            out_path = U'{}/{}.txt'.format(out_folder, filename)
            write_txt(out_path, head_rows + body_rows)
        else:
            raise ValueError('Output format is not support!')

        html_out_path = U'{}/{}.html'.format(html_out_folder, filename)
        print(U"\t\t-> '{}'".format(out_path))
        print(U"\t\t-> '{}'".format(html_out_path))

        open(html_out_path, 'wb').write(str(dom))

    pass


def load_xlsx_trans_file(xlsx_path):
    sheet_name = get_filename(xlsx_path, False)
    wb = load_workbook(xlsx_path, True)

    sheet = wb[sheet_name]

    rows_it = sheet.rows
    next(rows_it)  # skip header

    dict_result = {}

    for row in rows_it:
        line_id_cell = row[0]
        jp_text_cell = row[1]
        vn_text_cell = row[2]

        line_id = line_id_cell.value
        jp_text = jp_text_cell.value
        vn_text = vn_text_cell.value

        if line_id in dict_result:
            print(U'[WARNING]: Duplicate id: "{}" row={}'.format(line_id, line_id_cell.row))

        dict_result[line_id] = [jp_text, vn_text]
        pass

    return dict_result

    pass


def text_to_html(trans_folder, html_dump_folder, html_out_folder, in_format=XLSX_FORMAT):
    dict_format_support = {CSV_FORMAT: U'csv', TXT_FORMAT: U'txt', XLSX_FORMAT: U'xlsx'}
    f_format = dict_format_support.get(in_format)
    if f_format is None:
        raise 'Not support file format!'

    if not is_folder(html_out_folder):
        make_dirs(html_out_folder)

    trans_file_paths = glob.glob(U'{}/*.{}'.format(trans_folder, f_format))

    for trans_file_path in trans_file_paths:
        filename = get_filename(trans_file_path, False)
        html_out_path = U'{}/{}.html'.format(html_out_folder, filename)

        print(U"'{}' -> '{}'".format(trans_file_path, html_out_path))
        html_input_path = U'{}/{}.html'.format(html_dump_folder, filename)

        if not is_file(html_input_path):
            print(U'[WARNING]: File "{}" not found! -> skip'.format(html_input_path))
            continue

        html_text = open(html_input_path, 'rb').read()
        dom = BeautifulSoup(html_text, features="html.parser")

        dict_trans = {}
        if in_format == XLSX_FORMAT:
            dict_trans = load_xlsx_trans_file(trans_file_path)
        elif in_format == TXT_FORMAT:
            dict_trans = load_txt_trans_file(trans_file_path)

        head = dom.head  # type: Tag
        body = dom.body  # type: Tag

        trans_elems = body.find_all(TRANS_TAG)

        for line_id in dict_trans:
            jp_text, vn_text = dict_trans[line_id]
            trans_text = vn_text

            if trans_text is None or trans_text == U'':
                trans_text = jp_text

            line_id_clean = line_id.strip()
            if line_id_clean == TITLE_TAG:
                title_tag = head.find(TITLE_TAG)
                title_tag.string = trans_text
            elif line_id_clean.startswith(META_TAG):
                attr_val = line_id_clean.split(U'_', 1)[1]
                attr_name = None
                if attr_val in META_NAME:
                    attr_name = NAME
                elif attr_val in META_PROP:
                    attr_name = PROPERTY
                else:
                    print(U'[WARNING]: attr "{}" not found!'.format(attr_val))
                    continue

                elem = head.find(META_TAG, attrs={attr_name: attr_val})

                if elem is None:
                    print(U'[WARNING]: elem <meta {}="{}"> not found!'.format(attr_name, attr_val))
                    continue

                elem.attrs[CONTENT] = trans_text
            else:
                # elems = body.find_all(TRANS_TAG, attrs={'idx': line_id_clean})
                elems = []

                for elem in trans_elems:  # type: Tag
                    elem_idx = elem.attrs.get('idx')
                    if elem_idx is not None and elem_idx == line_id_clean:
                        elems.append(elem)

                if len(elems) == 0:
                    print(U'[WARNING]: elem <{} idx="{}"> not found!'.format(TRANS_TAG, line_id_clean))
                    continue

                for elem in elems:
                    new_elem = BeautifulSoup(trans_text, features="html.parser")
                    elem.replace_with(new_elem)

        open(html_out_path, 'wb').write(str(dom))

    pass


def main():
    # sys.argv.extend(['-t2h', 'data/trans', 'data/trans/html', 'data/out'])
    # sys.argv.extend(['-h2t', 'data/key.visualarts.gr.jp_full', 'data/dump'])

    parser = argparse.ArgumentParser(description='primaldoll tool')
    parser.add_argument('-v', '--version', action='version', version="1.0")

    parser.add_argument('-f', help='translate file format: txt=1 xlsx=2', type=int, default=TXT_FORMAT)

    parser.add_argument('-h2t', nargs=2, type=str, metavar=('html_folder', 'output_folder'),
                        help='html to text.')

    parser.add_argument('-t2h', nargs=3, type=str, metavar=('trans_folder', 'html_dump_folder', 'output_folder'),
                        help='text to html.')

    args = parser.parse_args()

    if len(sys.argv) <= 1:
        parser.print_help()

    if args.h2t:
        argv = args.h2t
        html_to_text(argv[0], argv[1], args.f)
    elif args.t2h:
        argv = args.t2h
        text_to_html(argv[0], argv[1], argv[2], args.f)

    pass


if __name__ == '__main__':
    main()
